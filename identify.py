import cognitive_face as CF
from global_variables import personGroupId
import os, urllib
import sqlite3
from openpyxl.cell import Cell 
from openpyxl.utils import get_column_letter, column_index_from_string
import time
from openpyxl import Workbook, load_workbook
import sys

#get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = "reports.xlsx")
sheet = wb.get_sheet_by_name('Cse15')

def getDateColumn():
	for i in range(1, len(sheet.rows[0]) + 1):
		col = get_column_letter(i)
		if sheet.cell('%s%s'% (col,'1')).value == currentDate:
			return col
			
			
Key = '302c3ee94b5a4a2896c7d2baecb035bb'
CF.Key.set(Key)
CF.util.DEFAULT_BASE_URL = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0/'
connect = sqlite3.connect("Face-DataBase")
c = connect.cursor()

attend = [0 for i in range(60)]	

currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, 'Cropped_faces')
for filename in os.listdir(directory):
	if filename.endswith(".jpg"):
		imgurl = urllib.pathname2url(os.path.join(directory, filename))
		res = CF.face.detect(imgurl)
		if len(res) != 1:
			print "No face detected."
			continue
			
		faceIds = []
		for face in res:
			faceIds.append(face['faceId'])
		res = CF.face.identify(faceIds, personGroupId)
		print filename
		print "Res is " + str(res)
		for face  in res:
			if not face['candidates']:
				print "Unknown"
			else:
				# personId = face['candidates'][0]['personId']
				personId = "81fda148"
				print "PersonId is : " + str(personId)
				c.execute("SELECT * FROM Students WHERE {0} = {1}".format("personId",personId))
				row = c.fetchone()
				print type(row)
				sys.exit(0)
				x = int(row[0])
				attend[x] += 1
				print row[1] + "recognized"
		time.sleep(6)
		
"""for row in range(2, len(sheet.columns[0]) + 1):
	rn = sheet.cell('A%s'% row).value
	if rn is not None:
		rn = rn[-2:]
		if attend[int(rn)] != 0:
			col = getDateColumn()
			sheet['%s%s' % (col, str(row))] = 1

wb.save(filename = "reports.xlsx")	 	
"""
for row in range(2, sheet.max_row + 1):
	rn = sheet.cell('A%s'% row,1).value
	if rn is not None:
		rn = rn[-2:]
		if attend[int(rn)] != 0:
			col = getDateColumn()
			sheet['%s%s' % (col, str(row))] = 1

wb.save("reports-out.xlsx")	 	

#currentDir = os.path.dirname(os.path.abspath(__file__))
#imgurl = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
#res = CF.face.detect(imgurl)
#faceIds = []
#for face in res:
 #   faceIds.append(face['faceId'])

#res = CF.face.identify(faceIds,personGroupId)
# for face in res:
#     personName = CF.person.get(personGroupId, face['candidates']['personId'])
#     print personName
#print res
